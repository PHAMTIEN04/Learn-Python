from tkinter import *
import openpyxl
import os
import datetime
import calendar

# Set the current working directory to the location of the Excel file
os.chdir(r"D:/Learn Python/Project/Tick Execl")

# Name of the Excel file
filename = 'TickLearn.xlsx'

# Function to update cell value in the Excel sheet
def tickexecl(filename, paper_sheet, row_col, value):
    workbook = openpyxl.load_workbook(filename=filename)
    sheet = workbook[paper_sheet]
    sheet[row_col].value = value
    workbook.close()
    workbook.save(filename=filename)

# Create the main window for the application
win = Tk()
win.geometry("665x550")
win.configure(bg="#F0F0F0")  # Set the background color of the main window
win.title("Attendance for Lessons")
win.iconbitmap("favicon.ico")

# Function to update the attendance data in the Excel sheet
def done():
    # Get the current date and time
    now = datetime.datetime.now()
    a = str(now).split("-")
    a[2] = a[2].split(" ")
    year = a[0]
    month = a[1]
    day = a[2][0]
    time = a[2][1]
    is_leap = calendar.isleap(int(year))

    # Update date and course names in the Excel sheet
    tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col="A1", value=month + "/" + day + "/" + year)
    tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col="A2", value="Cyber Security")
    tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col="A3", value="English B1")
    tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col="A4", value="Python Hacking")
    tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col="A5", value="Python Tkinter")
    tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col="AG1", value="Month Summary 7")

    # Determine the number of days in the current month
    while True:
        if month in ["01", "03", "05", "07", "08", "10", "12"]:
            number_of_day = 31
            break
        elif month in ["04", "06", "09", "11"]:
            number_of_day = 30
            break
        elif is_leap:
            number_of_day = 29
            break
        else:
            number_of_day = 28
            break

    cnt = 1
    alpha = "A"
    alpha_1 = "A"
    check_tick = ""
    while cnt <= number_of_day:
        wordbook1 = openpyxl.load_workbook(filename=filename)
        sheet = wordbook1["Trang_tính1"]
        run_check = ord(alpha)
        run_check1 = ord(alpha_1)

        if cnt < 26:
            run_check = run_check + 1
            alpha = chr(run_check)
            if cnt < 10:
                tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col=alpha + "1", value="0" + str(cnt))
                if sheet[alpha + "1"].value == day:
                    check_tick = alpha
            else:
                tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col=alpha + "1", value=str(cnt))
                if sheet[alpha + "1"].value == day:
                    check_tick = alpha
        elif cnt >= 26:
            tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col="A" + alpha_1 + "1", value=str(cnt))
            if sheet["A" + alpha_1 + "1"].value == day:
                check_tick = "A" + alpha_1
            run_check1 = run_check1 + 1
            alpha_1 = chr(run_check1)
        cnt = cnt + 1
        wordbook1.close()
        wordbook1.save(filename=filename)

    # Mark attendance for courses on the current day
    t = 2
    while t < 6:
        wordbook1 = openpyxl.load_workbook(filename=filename)
        sheet = wordbook1["Trang_tính1"]
        sheet[check_tick + str(t)].value = "✓"
        wordbook1.close()
        wordbook1.save(filename=filename)
        t = t + 1

    cnt1 = 1
    alpha_check = "A"
    alpha_check1 = "A"
    cnt_c = 0
    cnt_pt = 0
    cnt_ph = 0
    cnt_eng = 0

    # Calculate the number of days for each course
    while cnt1 <= number_of_day:
        wordbook1 = openpyxl.load_workbook(filename=filename)
        sheet = wordbook1["Trang_tính1"]
        run_check_a = ord(alpha_check)
        run_check_b = ord(alpha_check1)
        if cnt1 < 26:
            run_check_a = run_check_a + 1
            alpha_check = chr(run_check_a)
            if sheet[alpha_check + "2"].value == "✓":
                cnt_c = cnt_c + 1
            if sheet[alpha_check + "3"].value == "✓":
                cnt_pt = cnt_pt + 1
            if sheet[alpha_check + "4"].value == "✓":
                cnt_ph = cnt_ph + 1
            if sheet[alpha_check + "5"].value == "✓":
                cnt_eng = cnt_eng + 1
        elif cnt1 >= 26:
            if sheet["A" + alpha_check1 + "2"].value == "✓":
                cnt_c = cnt_c + 1
            if sheet["A" + alpha_check1 + "3"].value == "✓":
                cnt_pt = cnt_pt + 1
            if sheet["A" + alpha_check1 + "4"].value == "✓":
                cnt_ph = cnt_ph + 1
            if sheet["A" + alpha_check1 + "5"].value == "✓":
                cnt_eng = cnt_eng + 1
            run_check_b = run_check_b + 1
            alpha_check1 = chr(run_check_b)
        cnt1 = cnt1 + 1
        wordbook1.close()
        wordbook1.save(filename=filename)

    # Update the summary data for each course and the total number of days attended
    t_sum = 2
    list_t = [cnt_c, cnt_ph, cnt_pt, cnt_eng]
    i = 0

    while t_sum < 7:
        if t_sum < 6:
            tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col="AG" + str(t_sum), value="Finish " + str(list_t[i]) + " Day")
            i = i + 1
        else:
            tickexecl(filename=filename, paper_sheet="Trang_tính1", row_col="AG" + str(t_sum), value="Total: " + str(sum(list_t)))
        t_sum = t_sum + 1

    # Update the labels with course completion data
    lb_text = Label(lb_f1, text="Success Cyber Security", bg="#009688", fg="white", font=("Arial", 14, "bold"))
    lb_text.grid(column=0, row=0, padx=10, pady=10, sticky="W")

    lb_text1 = Label(lb_f1, text="Success English B1", bg="#FFC107", fg="black", font=("Arial", 14, "bold"))
    lb_text1.grid(column=0, row=1, padx=10, pady=10, sticky="W")

    lb_text2 = Label(lb_f1, text="Success Python Hacking", bg="#FF5722", fg="white", font=("Arial", 14, "bold"))
    lb_text2.grid(column=0, row=2, padx=10, pady=10, sticky="W")

    lb_text3 = Label(lb_f1, text="Success Python Tkinter", bg="#FF5722", fg="white", font=("Arial", 14, "bold"))
    lb_text3.grid(column=0, row=3, padx=10, pady=10, sticky="W")

# Create label frames for layout
lb_f = LabelFrame(win, pady=50, bg="#F0F0F0")  # Set the background color of the label frame
lb_f.grid(padx=10, pady=10)

lb_f1 = LabelFrame(lb_f, bg="#F0F0F0")  # Set the background color of the nested label frame
lb_f1.grid(row=1, rowspan=2, column=6, padx=10, columnspan=10, sticky="nsew")

lb_none = Label(lb_f1, text="", bg="#F0F0F0")  # Set the background color of the label
lb_none.grid(columnspan=6, column=3, row=0)

lb = Label(lb_f, text="", padx=200, bg="#F0F0F0")  # Set the background color of the label
lb.grid(columnspan=10, column=5, row=0)

# Create buttons
bt1 = Button(lb_f, text="Success", padx=50, pady=50, command=done, bg="#2196F3", fg="white", font=("Arial", 14, "bold"), border=3)  # Set the colors of the button
bt1.grid(column=0, row=1, padx=20, pady=10)

bt2 = Button(lb_f, text="Quit", padx=50, pady=50, command=win.quit, bg="#F44336", fg="white", font=("Arial", 14, "bold"), border=3)  # Set the colors of the button
bt2.grid(column=0, row=2, padx=20, pady=50)

win.mainloop()
