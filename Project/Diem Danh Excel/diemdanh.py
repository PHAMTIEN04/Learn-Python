import tkinter as tk
import openpyxl
import os

os.chdir(r"D:/Learn Python/Project/Diem Danh Excel")
filename = "diemdanh.xlsx"
alpha = "A"
check = 0
i = 1
list_sv = ["Phạm Phước Tiến","Lê Vũ Trang","Nguyễn Thanh An","Phạm Nam Giao","Lê Vũ An","Trần Thị Thúy","Nguyễn An Nhiên","Bùi Xuân Hiếu","Trần Ngọc Thảo"]
def tickexecl(filename, paper_sheet, row_col, value):
    workbook = openpyxl.load_workbook(filename=filename)
    sheet = workbook[paper_sheet]
    sheet[row_col].value = value
    workbook.save(filename=filename)
    workbook.close()

def khongvang():
    global check
    global alpha 
    worbook = openpyxl.load_workbook(filename=filename)
    sheet = worbook["Trang_tính1"]
    
    check  = int(sheet["A1"].value)
    worbook.close()
    worbook.save(filename=filename)
    cnt = 2
    
    if check < 6:
        run_check = ord(alpha)
        run_check = run_check + check
        alpha = chr(run_check)
        while cnt < len(list_sv)+2 :
            tickexecl(filename=filename,paper_sheet="Trang_tính1",row_col=alpha + str(cnt),value="✓")
            cnt = cnt+1    

        worbook = openpyxl.load_workbook(filename=filename)
        sheet = worbook["Trang_tính1"]
        sheet["A1"].value = check + 1
        worbook.close()
        worbook.save(filename=filename)

def vang():
    global i
    global check
    global alpha 
    check_get = lb_e.get()

    cnt1 = 2
    check_bool = False
    while cnt1 < len(list_sv)+2:
        worbook = openpyxl.load_workbook(filename=filename)
        sheet = worbook["Trang_tính1"]
        if sheet["A"+str(cnt1)].value == check_get:
            tickexecl(filename=filename,paper_sheet="Trang_tính1",row_col=alpha+str(cnt1),value="X")
            check_bool = True
            break
        
        cnt1 = cnt1 + 1
        worbook.close()
        worbook.save(filename=filename)
    if check_bool == True:
        lb_v = tk.Label(lb_f1,text=check_get,bg="#FF5722", fg="white", font=("Arial", 10, "bold"))
        lb_v.grid(columnspan=1,column=0,row=i,padx=5,pady=5)
        i = i + 1

win = tk.Tk()
win.geometry("870x450")
win.title("Điểm Danh")
win["bg"] = "#D3D3D3"
win.iconbitmap("favicon.ico")
lb_f = tk.LabelFrame(win,bg="#D3D3D3")
lb_f.grid(padx=10,pady=20)

lb_f1 = tk.LabelFrame(lb_f,bg="#D3D3D3")
lb_f1.grid(rowspan=3,columnspan= 9,column=4,row=1,sticky="nsew")

lb_n = tk.Label(lb_f1,text="Danh Sách Vắng :",bg="#D3D3D3",fg="black",font=("Arial", 12, "bold"))
lb_n.grid(columnspan=6,column=0,row=0)

lb_f2 = tk.LabelFrame(lb_f,bg="#D3D3D3")
lb_f2.grid(rowspan=3,columnspan= 12,column=13,row=1,sticky="nsew")

lb_n1 = tk.Label(lb_f2,text="Danh Sách Sinh Viên :",bg="#D3D3D3",fg="black",font=("Arial", 12, "bold"))
lb_n1.grid(columnspan=6,column=0,row=0)
run_i = 0
while run_i < len(list_sv):
    lb_sv = tk.Label(lb_f2,text=str(run_i+1)+"."+str(list_sv[run_i]),bg="#D3D3D3",font=("Arial", 10, "bold"))
    lb_sv.grid(column=0,row=run_i+1)
    run_i = run_i + 1

lb = tk.Label(lb_f,text="",width=120,bg="#D3D3D3")
lb.grid(columnspan=21,row=0,column=0)

lb_e = tk.Entry(lb_f,width=30,border=5)
lb_e.grid(columnspan=2,row=1,column=0,padx=10)


bt_e = tk.Button(lb_f,text="Vắng",border=3,width=5,command=vang,bg="#F44336", fg="white", font=("Arial", 8, "bold"))
bt_e.grid(columnspan=5,column=1,row=1)

bt_all = tk.Button(lb_f,text="Điểm Danh",width=25,height=5,command=khongvang, bg="#2196F3", fg="white", font=("Arial", 9, "bold"),border=3)
bt_all.grid(columnspan=2,row=2,column=0,pady = 80)

bt_q = tk.Button(lb_f,text="Thoát",width=25,height=5,command=win.quit,bg="#2196F3", fg="white", font=("Arial", 9, "bold"),border=3)
bt_q.grid(columnspan=2,row=3,column=0,pady=10)

win.mainloop()