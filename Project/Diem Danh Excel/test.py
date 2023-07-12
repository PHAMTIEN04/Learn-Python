import openpyxl
import os
import calendar
import datetime 
os.chdir(r"D:/Learn Python/Project/Diem Danh Excel")

filename = "diemdanh.xlsx"

def tickexecl(filename, paper_sheet, row_col, value):
    workbook = openpyxl.load_workbook(filename=filename)
    sheet = workbook[paper_sheet]
    sheet[row_col].value = value
    workbook.close()
    workbook.save(filename=filename)

worbook = openpyxl.load_workbook(filename=filename)
sheet = worbook["Trang_tính1"]
list_sv = ["Phạm Phước Tiến","Lê Vũ Trang","Nguyễn Thanh An"]
check  = int(sheet["A1"].value)
worbook.close()
worbook.save(filename=filename)
cnt = 2
alpha = "A"
if check < 6:
    run_check = ord(alpha)
    run_check = run_check + check
    alpha = chr(run_check)
    check_t = False
    while cnt < len(list_sv)+2 :
        check_t = True
        tickexecl(filename=filename,paper_sheet="Trang_tính1",row_col=alpha + str(cnt),value="✓")
        cnt = cnt+1    

    worbook = openpyxl.load_workbook(filename=filename)
    sheet = worbook["Trang_tính1"]
    sheet["A1"].value = check + 1
    worbook.close()
    worbook.save(filename=filename)
    
check = str(input())
cnt1 = 2
while True:
    worbook = openpyxl.load_workbook(filename=filename)
    sheet = worbook["Trang_tính1"]
    if sheet3["A"+str(cnt1)].value == check:
        tickexecl(filename=filename,paper_sheet="Trang_tính1",row_col=alpha+str(cnt1),value="X")
        break
    cnt1 = cnt1 + 1
    worbook3.close()
    worbook3.save(filename=filename)


